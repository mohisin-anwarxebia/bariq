"""
Live Data Import API — Upload Excel/JSON to update demand forecast data in real-time.

Supported Excel format (demand_forecast.xlsx):
  Column A: category (e.g., Beer, Cocktails, Tequila, Chicken Wings, Food Overall)
  Column B: change_pct (e.g., 18, 24, 27, 21, 19)
  Column C: confidence (e.g., 0.86) — optional
  Column D: factor_type (e.g., weather, event, historical) — optional
  Column E: factor_description (e.g., "Temperature: 94°F") — optional
  Column F: factor_impact (e.g., "+5% beverage demand") — optional

Or upload with a "forecasts" sheet and a "factors" sheet.

API also accepts JSON POST for programmatic updates.
"""
from fastapi import APIRouter, UploadFile, File, Depends
from pydantic import BaseModel
from typing import Optional, List
from sqlalchemy.orm import Session
from app.database import get_db
import json
import os

router = APIRouter()

# In-memory live forecast data (overrides default when set)
_live_forecast_data: Optional[dict] = None


def get_live_forecast():
    """Return live forecast data if uploaded, else None."""
    return _live_forecast_data


def clear_live_forecast():
    """Clear live forecast data, reverting to defaults."""
    global _live_forecast_data
    _live_forecast_data = None


class ForecastItem(BaseModel):
    category: str
    change_pct: float


class ForecastFactor(BaseModel):
    type: str
    description: str
    impact: str
    distance: Optional[str] = None


class LiveForecastData(BaseModel):
    forecast_date: Optional[str] = None
    day: Optional[str] = None
    forecasts: List[ForecastItem]
    factors: Optional[List[ForecastFactor]] = []
    confidence: Optional[float] = 0.86
    inventory_risks: Optional[List[dict]] = []


@router.post("/data/forecast/upload")
async def upload_forecast_excel(file: UploadFile = File(...)):
    """Upload an Excel file to update demand forecast data live."""
    global _live_forecast_data

    if not file.filename.endswith(('.xlsx', '.xls')):
        return {"error": "File must be .xlsx or .xls format"}

    try:
        import openpyxl
        from io import BytesIO
        from datetime import datetime, timedelta

        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True)

        forecasts = []
        factors = []
        confidence = 0.86
        inventory_risks = []

        # Try "forecasts" sheet first, fall back to first sheet
        if "forecasts" in wb.sheetnames:
            ws = wb["forecasts"]
        else:
            ws = wb.active

        # Read forecasts: expect header row then data
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if row and row[0] and row[1] is not None:
                category = str(row[0]).strip()
                change_pct = float(row[1])
                forecasts.append({"category": category, "change_pct": change_pct})
                if len(row) > 2 and row[2]:
                    confidence = float(row[2])

        # Try "factors" sheet
        if "factors" in wb.sheetnames:
            ws_factors = wb["factors"]
            factor_rows = list(ws_factors.iter_rows(min_row=2, values_only=True))
            for row in factor_rows:
                if row and row[0]:
                    factors.append({
                        "type": str(row[0]).strip(),
                        "description": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                        "impact": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                        "distance": str(row[3]).strip() if len(row) > 3 and row[3] else None
                    })

        # Try "risks" sheet
        if "risks" in wb.sheetnames:
            ws_risks = wb["risks"]
            risk_rows = list(ws_risks.iter_rows(min_row=2, values_only=True))
            for row in risk_rows:
                if row and row[0]:
                    inventory_risks.append({
                        "product": str(row[0]).strip(),
                        "current_stock": str(row[1]) if len(row) > 1 and row[1] else "",
                        "forecast_demand": str(row[2]) if len(row) > 2 and row[2] else "",
                        "risk": str(row[3]).strip() if len(row) > 3 and row[3] else "Stock-out risk",
                        "recommendation": str(row[4]).strip() if len(row) > 4 and row[4] else ""
                    })

        # If factors not in separate sheet, check columns D/E/F in main sheet
        if not factors:
            for row in rows:
                if row and len(row) > 3 and row[3]:
                    factors.append({
                        "type": str(row[3]).strip(),
                        "description": str(row[4]).strip() if len(row) > 4 and row[4] else "",
                        "impact": str(row[5]).strip() if len(row) > 5 and row[5] else ""
                    })

        if not forecasts:
            return {"error": "No forecast data found in the Excel file. Expected columns: category, change_pct"}

        # Calculate forecast date (next Saturday)
        today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        saturday = today + timedelta(days=(5 - today.weekday()) % 7)

        _live_forecast_data = {
            "forecast_date": str(saturday.date()),
            "day": "Saturday",
            "forecasts": forecasts,
            "factors": factors if factors else [
                {"type": "uploaded", "description": "Data from uploaded Excel file", "impact": "Custom forecast"}
            ],
            "confidence": confidence,
            "inventory_risks": inventory_risks
        }

        wb.close()

        return {
            "status": "success",
            "message": f"Forecast updated with {len(forecasts)} categories and {len(factors)} factors",
            "data": _live_forecast_data
        }

    except Exception as e:
        return {"error": f"Failed to parse Excel file: {str(e)}"}


@router.post("/data/forecast/json")
async def update_forecast_json(data: LiveForecastData):
    """Update demand forecast data via JSON API."""
    global _live_forecast_data
    from datetime import datetime, timedelta

    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    saturday = today + timedelta(days=(5 - today.weekday()) % 7)

    _live_forecast_data = {
        "forecast_date": data.forecast_date or str(saturday.date()),
        "day": data.day or "Saturday",
        "forecasts": [f.dict() for f in data.forecasts],
        "factors": [f.dict() for f in (data.factors or [])],
        "confidence": data.confidence or 0.86,
        "inventory_risks": data.inventory_risks or []
    }

    return {
        "status": "success",
        "message": f"Forecast updated with {len(data.forecasts)} categories",
        "data": _live_forecast_data
    }


@router.delete("/data/forecast")
async def clear_forecast():
    """Clear uploaded forecast data, revert to default demo data."""
    clear_live_forecast()
    return {"status": "success", "message": "Live forecast data cleared. Reverted to demo defaults."}


@router.get("/data/forecast/status")
async def forecast_status():
    """Check if live forecast data is active."""
    if _live_forecast_data:
        return {
            "source": "live_upload",
            "categories": len(_live_forecast_data["forecasts"]),
            "factors": len(_live_forecast_data["factors"]),
            "data": _live_forecast_data
        }
    return {"source": "demo_default", "message": "Using seeded demo data. Upload Excel or POST JSON to override."}


@router.get("/data/forecast/template")
async def download_template_info():
    """Return the expected Excel template format."""
    return {
        "template": {
            "sheet_1_name": "forecasts (or use the first/active sheet)",
            "columns": {
                "A": "category — e.g., Beer, Cocktails, Tequila, Chicken Wings, Food Overall",
                "B": "change_pct — e.g., 18, 24, 27 (percentage increase/decrease)",
                "C": "confidence — optional, e.g., 0.86 (0 to 1)"
            },
            "sheet_2_name": "factors (optional)",
            "factor_columns": {
                "A": "type — weather, event, historical",
                "B": "description — e.g., Temperature: 94°F, Sunny",
                "C": "impact — e.g., +5% beverage demand",
                "D": "distance — optional, e.g., 1.2 miles away"
            },
            "sheet_3_name": "risks (optional)",
            "risk_columns": {
                "A": "product — e.g., Don Julio 1942",
                "B": "current_stock — e.g., 16.3",
                "C": "forecast_demand — e.g., 22",
                "D": "risk — e.g., Stock-out risk",
                "E": "recommendation — e.g., Order 4 cases"
            }
        },
        "json_api": {
            "endpoint": "POST /api/data/forecast/json",
            "example": {
                "forecasts": [
                    {"category": "Beer", "change_pct": 18},
                    {"category": "Cocktails", "change_pct": 24},
                    {"category": "Tequila", "change_pct": 27}
                ],
                "factors": [
                    {"type": "weather", "description": "Temperature: 94°F", "impact": "+5% beverage demand"}
                ],
                "confidence": 0.86
            }
        }
    }
